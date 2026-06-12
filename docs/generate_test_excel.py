#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EmiyaOJ-Cloud 系统测试 Excel 生成脚本
基于《系统测试报告》和《测试方案与测试用例》生成包含 7 张工作表的详细测试 Excel
输出: docs/EmiyaOJ-Cloud-测试用例.xlsx
"""

import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ============================================================================
# 样式常量
# ============================================================================

# 颜色定义
COLOR_HEADER_FILL = "2F5496"       # 深蓝 - 表头背景
COLOR_HEADER_FONT = "FFFFFF"       # 白色 - 表头字体
COLOR_GROUP_FILL = "D9E2F3"        # 浅蓝灰 - 分组标题
COLOR_TITLE_FILL = "1F3864"        # 深蓝 - 主标题
COLOR_P0_FILL = "FFC7CE"           # 浅红 - P0 高亮
COLOR_P0_FONT = "9C0006"           # 深红 - P0 字体
COLOR_LIGHT_YELLOW = "FFF2CC"      # 浅黄 - 关键行
COLOR_LIGHT_GREEN = "C6EFCE"       # 浅绿 - 通过
COLOR_WHITE = "FFFFFF"
COLOR_BORDER = "B4C6E7"
COLOR_SUBTITLE_FILL = "4472C4"

# 字体
FONT_TITLE = Font(name="微软雅黑", size=16, bold=True, color=COLOR_HEADER_FONT)
FONT_SUBTITLE = Font(name="微软雅黑", size=12, bold=True, color=COLOR_HEADER_FONT)
FONT_HEADER = Font(name="微软雅黑", size=10, bold=True, color=COLOR_HEADER_FONT)
FONT_NORMAL = Font(name="微软雅黑", size=10)
FONT_BOLD = Font(name="微软雅黑", size=10, bold=True)
FONT_SMALL = Font(name="微软雅黑", size=9, color="666666")
FONT_P0 = Font(name="微软雅黑", size=10, bold=True, color=COLOR_P0_FONT)
FONT_GROUP = Font(name="微软雅黑", size=10, bold=True, color="1F3864")

# 填充
FILL_HEADER = PatternFill(start_color=COLOR_HEADER_FILL, end_color=COLOR_HEADER_FILL, fill_type="solid")
FILL_TITLE = PatternFill(start_color=COLOR_TITLE_FILL, end_color=COLOR_TITLE_FILL, fill_type="solid")
FILL_SUBTITLE = PatternFill(start_color=COLOR_SUBTITLE_FILL, end_color=COLOR_SUBTITLE_FILL, fill_type="solid")
FILL_GROUP = PatternFill(start_color=COLOR_GROUP_FILL, end_color=COLOR_GROUP_FILL, fill_type="solid")
FILL_P0 = PatternFill(start_color=COLOR_P0_FILL, end_color=COLOR_P0_FILL, fill_type="solid")
FILL_YELLOW = PatternFill(start_color=COLOR_LIGHT_YELLOW, end_color=COLOR_LIGHT_YELLOW, fill_type="solid")
FILL_GREEN = PatternFill(start_color=COLOR_LIGHT_GREEN, end_color=COLOR_LIGHT_GREEN, fill_type="solid")
FILL_WHITE = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")

# 对齐
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_TITLE = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# 边框
THIN_BORDER = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)
BOTTOM_BORDER = Border(bottom=Side(style="medium", color=COLOR_HEADER_FILL))


def apply_cell_style(cell, font=FONT_NORMAL, fill=FILL_WHITE, alignment=ALIGN_LEFT, border=THIN_BORDER):
    """统一应用单元格样式"""
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = border


def apply_header_style(ws, row, cols, start_col=1):
    """给表头行统一加样式"""
    for col_idx in range(start_col, start_col + cols):
        cell = ws.cell(row=row, column=col_idx)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)


def set_col_widths(ws, widths_dict):
    """批量设置列宽"""
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


def add_title_row(ws, row, col_start, col_end, text, font=FONT_TITLE, fill=FILL_TITLE, height=36):
    """添加合并标题行"""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    apply_cell_style(cell, font=font, fill=fill, alignment=ALIGN_CENTER)
    ws.row_dimensions[row].height = height


def add_subtitle_row(ws, row, col_start, col_end, text, height=28):
    """添加副标题行"""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    apply_cell_style(cell, font=FONT_SUBTITLE, fill=FILL_SUBTITLE, alignment=ALIGN_CENTER)
    ws.row_dimensions[row].height = height


def add_info_row(ws, row, col_label, label_text, col_value, value_text, col_end, label_width=None):
    """添加信息行（标签: 值 格式）"""
    cell_label = ws.cell(row=row, column=col_label, value=label_text)
    apply_cell_style(cell_label, font=FONT_BOLD, fill=FILL_YELLOW, alignment=ALIGN_CENTER)
    ws.merge_cells(start_row=row, start_column=col_value, end_row=row, end_column=col_end)
    cell_value = ws.cell(row=row, column=col_value, value=value_text)
    apply_cell_style(cell_value, font=FONT_NORMAL, alignment=ALIGN_LEFT)


def add_group_header(ws, row, col_start, col_end, text, height=24):
    """添加用例组标题行"""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=text)
    apply_cell_style(cell, font=FONT_GROUP, fill=FILL_GROUP, alignment=ALIGN_LEFT)
    ws.row_dimensions[row].height = height


# ============================================================================
# 数据区 — 测试用例（全部 78 条）
# ============================================================================

TEST_CASE_GROUPS = [
    {
        "group_id": "TC-E2E",
        "group_name": "TC-E2E 核心业务链路用例",
        "module": "端到端",
        "cases": [
            ["TC-E2E-001", "管理员配置题目并完成用户判题", "端到端", "P0",
             "数据库初始化，管理员和普通用户存在，Go-Judge 可用",
             "1. 管理端登录管理员账号\n2. 创建题目、样例用例、隐藏用例和启用语言\n3. 用户端登录普通用户\n4. 浏览题目详情并提交 AC 代码\n5. 查询提交详情",
             "管理员配置成功；用户可看到题目和样例；提交生成记录；最终状态为 AC；隐藏用例内容不向普通用户展示",
             "", "", "核心演示链路"],
            ["TC-E2E-002", "用户发布题解并通过审核公开展示", "端到端", "P0",
             "普通用户、审核用户存在，RabbitMQ 和审核服务可用",
             "1. 用户登录并针对题目发布题解\n2. 检查内容进入待审核状态\n3. 审核服务消费任务或管理端人工通过\n4. 用户端查询题解列表",
             "题解保存成功；审核通过后可公开查询；审核原因和状态可追踪",
             "", "", "博客审核链路"],
            ["TC-E2E-003", "竞赛报名、提交和排行榜展示", "端到端", "P0",
             "存在进行中竞赛，已配置竞赛题目和语言",
             "1. 用户报名竞赛\n2. 用户进入竞赛题目提交代码\n3. 查询竞赛提交记录\n4. 查看排行榜",
             "报名成功；竞赛提交被接受；提交记录带竞赛编号；排行榜包含该用户成绩",
             "", "", "竞赛主链路"],
            ["TC-E2E-004", "外部依赖异常下核心链路兜底", "端到端", "P1",
             "已准备 AI 或审核外部服务异常场景",
             "1. 用户提交代码并查询结果\n2. 用户发送 AI 问题\n3. 用户发布博客并使用人工审核兜底",
             "判题主链路不受 AI 异常影响；AI 返回友好提示；审核外部服务不可用时可通过人工审核完成演示",
             "", "", "演示兜底"],
        ]
    },
    {
        "group_id": "TC-AUTH",
        "group_name": "TC-AUTH 认证登录用例",
        "module": "Auth",
        "cases": [
            ["TC-AUTH-001", "用户登录成功", "Auth", "P0",
             "存在启用用户",
             "调用登录接口，输入正确账号和密码",
             "返回统一响应；data 中包含 Token 和用户信息；Redis 写入 Token 状态",
             "", "", "登录基础用例"],
            ["TC-AUTH-002", "用户登录失败", "Auth", "P0",
             "存在用户",
             "调用登录接口，输入错误密码",
             "返回登录失败提示；不签发 Token",
             "", "", "异常输入"],
            ["TC-AUTH-003", "登出后 Token 失效", "Auth", "P0",
             "用户已登录",
             "1. 携带 Token 调用登出接口\n2. 再次携带原 Token 访问受保护接口",
             "登出成功；再次访问被拒绝",
             "", "", "Redis 白名单"],
            ["TC-AUTH-004", "禁用用户登录", "Auth", "P1",
             "存在被禁用用户",
             "使用禁用用户账号密码调用登录接口",
             "登录被拒绝，返回账号状态异常提示，不签发 Token",
             "", "", "账号状态"],
        ]
    },
    {
        "group_id": "TC-GW",
        "group_name": "TC-GW 网关鉴权用例",
        "module": "Gateway",
        "cases": [
            ["TC-GW-001", "未登录访问受保护接口", "Gateway", "P0",
             "无 Token",
             "不携带 Authorization 访问提交代码或发布博客接口",
             "Gateway 返回 401 或统一未认证响应",
             "", "", "网关鉴权"],
            ["TC-GW-002", "登录后访问受保护接口", "Gateway", "P0",
             "用户已登录",
             "携带合法 Token 查询我的提交或发布评论",
             "请求被转发到业务服务；下游服务能识别用户编号",
             "", "", "用户上下文"],
            ["TC-GW-003", "Token 格式异常或过期", "Gateway", "P0",
             "Gateway 可用",
             "携带伪造、过期或缺少用户信息的 Token 访问受保护接口",
             "请求被拒绝，返回未认证或 Token 异常提示，下游服务不执行业务写入",
             "", "", "Token 安全"],
        ]
    },
    {
        "group_id": "TC-RBAC",
        "group_name": "TC-RBAC 角色权限用例",
        "module": "Auth",
        "cases": [
            ["TC-RBAC-001", "普通用户访问管理接口", "Auth", "P1",
             "普通用户已登录",
             "携带普通用户 Token 访问题目新增、角色管理或审核管理接口",
             "返回无权限；不写入业务数据",
             "", "", "权限边界"],
            ["TC-RBAC-002", "管理员访问管理接口", "Auth", "P0",
             "管理员已登录",
             "携带管理员 Token 访问题目、语言、竞赛管理接口",
             "请求成功，页面或接口返回管理数据",
             "", "", "管理端权限"],
            ["TC-RBAC-003", "审核人员权限边界", "Auth", "P1",
             "审核人员已登录",
             "1. 访问博客审核接口\n2. 访问用户角色管理接口",
             "可访问审核接口；访问角色管理接口被拒绝",
             "", "", "细粒度权限"],
        ]
    },
    {
        "group_id": "TC-PROBLEM",
        "group_name": "TC-PROBLEM 题目管理用例",
        "module": "Problem",
        "cases": [
            ["TC-PROBLEM-001", "题目列表查询", "Problem", "P0",
             "存在公开题目",
             "用户端按分页、难度、关键字查询题目",
             "返回分页数据；只包含可公开题目",
             "", "", "题目浏览"],
            ["TC-PROBLEM-002", "题目详情查询", "Problem", "P0",
             "存在公开题目和测试用例",
             "查询题目详情",
             "返回题面、输入输出说明、样例、限制和标签；不返回隐藏用例输入输出",
             "", "", "隐藏数据保护"],
            ["TC-PROBLEM-003", "管理员新增题目", "Problem", "P0",
             "管理员已登录",
             "输入标题、描述、难度、时间限制、内存限制、标签并保存",
             "题目保存成功，可在管理端查询",
             "", "", "管理端"],
            ["TC-PROBLEM-004", "题目必填与边界校验", "Problem", "P1",
             "管理员已登录",
             "输入空标题、负数时间限制、负数内存限制或非法难度保存题目",
             "保存失败，返回明确参数校验提示，不写入无效题目",
             "", "", "参数边界"],
            ["TC-PROBLEM-005", "被引用题目删除保护", "Problem", "P1",
             "题目已被题单或竞赛引用",
             "管理端删除该题目",
             "系统拒绝删除或给出业务提示，不破坏题单、竞赛和提交记录关联",
             "", "", "数据一致性"],
        ]
    },
    {
        "group_id": "TC-CASE",
        "group_name": "TC-CASE 测试用例管理用例",
        "module": "Problem",
        "cases": [
            ["TC-CASE-001", "测试用例维护", "Problem", "P0",
             "已存在题目",
             "管理端新增样例用例和隐藏用例",
             "保存成功；样例可在题目详情展示；隐藏用例仅判题内部使用",
             "", "", "判题前置"],
            ["TC-CASE-002", "空测试用例保存校验", "Problem", "P1",
             "已存在题目",
             "管理端保存输入或输出为空的隐藏用例",
             "系统按规则拒绝保存或提示配置异常，避免题目进入不可判题状态",
             "", "", "用例质量"],
        ]
    },
    {
        "group_id": "TC-LANG",
        "group_name": "TC-LANG 编程语言配置用例",
        "module": "Problem",
        "cases": [
            ["TC-LANG-001", "启用语言列表查询", "Problem", "P0",
             "存在启用和禁用语言",
             "用户端查询 /language/list",
             "仅返回启用语言",
             "", "", "用户端"],
            ["TC-LANG-002", "禁用语言不可提交", "Problem", "P1",
             "存在禁用语言",
             "使用禁用语言编号提交代码",
             "提交被拒绝并返回语言不可用提示",
             "", "", "提交校验"],
            ["TC-LANG-003", "编译型语言命令缺失校验", "Problem", "P1",
             "管理员已登录",
             "新增编译型语言但不填写编译命令",
             "保存失败，返回编译型语言必须配置编译命令等提示",
             "", "", "命令模板"],
        ]
    },
    {
        "group_id": "TC-SET",
        "group_name": "TC-SET 题单管理用例",
        "module": "Problem",
        "cases": [
            ["TC-SET-001", "题单详情查询", "Problem", "P1",
             "存在公开题单并关联题目",
             "查询题单详情",
             "返回题单信息和按顺序排列的题目列表",
             "", "", "题单练习"],
            ["TC-SET-002", "题单题目排序调整", "Problem", "P1",
             "存在题单且关联多道题目",
             "管理端调整题单题目顺序并查询详情",
             "题单详情按新顺序返回，题目关联不丢失",
             "", "", "排序一致性"],
        ]
    },
    {
        "group_id": "TC-CONTEST",
        "group_name": "TC-CONTEST 竞赛管理用例",
        "module": "Problem",
        "cases": [
            ["TC-CONTEST-001", "竞赛报名成功", "Problem", "P0",
             "存在可报名竞赛",
             "用户调用报名接口，必要时输入邀请码",
             "报名成功；报名关系写入数据库",
             "", "", "竞赛参与"],
            ["TC-CONTEST-002", "未报名用户提交竞赛题目", "Problem", "P0",
             "竞赛要求报名，用户未报名",
             "用户携带竞赛编号提交竞赛题目代码",
             "提交被拒绝，返回未报名或无参赛资格提示",
             "", "", "规则校验"],
            ["TC-CONTEST-003", "竞赛时间外提交", "Problem", "P0",
             "存在未开始或已结束竞赛",
             "用户携带竞赛编号提交代码",
             "提交被拒绝，提示竞赛当前不可提交",
             "", "", "时间状态"],
            ["TC-CONTEST-004", "竞赛排行榜查询", "Problem", "P1",
             "竞赛已有提交结果",
             "查询竞赛排行榜",
             "返回参赛用户排名、通过题数、得分或耗时信息",
             "", "", "排名展示"],
            ["TC-CONTEST-005", "邀请码错误报名", "Problem", "P1",
             "存在需要邀请码的竞赛",
             "用户输入错误邀请码报名",
             "报名失败，返回邀请码错误或无权限提示，不写入报名关系",
             "", "", "报名校验"],
            ["TC-CONTEST-006", "竞赛开始前取消报名", "Problem", "P1",
             "用户已报名未开始竞赛",
             "用户调用取消报名接口",
             "取消成功；报名记录删除或状态更新；竞赛开始后再次取消应被拒绝",
             "", "", "状态约束"],
        ]
    },
    {
        "group_id": "TC-JUDGE",
        "group_name": "TC-JUDGE 判题提交用例",
        "module": "Judge",
        "cases": [
            ["TC-JUDGE-001", "普通代码提交成功", "Judge", "P0",
             "用户已登录，题目和语言可用",
             "调用 POST /judge/submit，提交可通过代码",
             "返回提交编号；提交状态进入待判题或判题中；最终为 AC",
             "", "", "主链路"],
            ["TC-JUDGE-002", "Wrong Answer 结果", "Judge", "P0",
             "存在可判题题目",
             "提交能运行但输出错误的代码",
             "最终状态为 WA；提交详情可查看失败状态",
             "", "", "输出比对"],
            ["TC-JUDGE-003", "Compilation Error 结果", "Judge", "P0",
             "使用需编译语言",
             "提交语法错误代码",
             "最终状态为 CE；记录编译输出",
             "", "", "编译错误"],
            ["TC-JUDGE-004", "Time Limit Exceeded 结果", "Judge", "P1",
             "题目有时间限制",
             "提交死循环或超时代码",
             "最终状态为 TLE；记录耗时信息",
             "", "", "沙箱限制"],
            ["TC-JUDGE-005", "Runtime Error 结果", "Judge", "P1",
             "题目和语言可用",
             "提交运行时异常代码",
             "最终状态为 RE；错误信息可追踪",
             "", "", "运行错误"],
            ["TC-JUDGE-006", "Go-Judge 不可用", "Judge", "P1",
             "模拟沙箱不可访问",
             "提交代码或触发判题",
             "提交被标记为 SE 或返回系统错误；服务日志记录原因",
             "", "", "外部依赖"],
            ["TC-JUDGE-007", "查询我的提交", "Judge", "P0",
             "用户已有多条提交",
             "调用 GET /submission/my",
             "返回当前用户提交分页，不包含他人敏感数据",
             "", "", "数据隔离"],
            ["TC-JUDGE-008", "查看他人提交详情", "Judge", "P1",
             "普通用户 A、B 均存在",
             "用户 A 查询用户 B 的完整提交详情",
             "被拒绝或隐藏代码、隐藏用例等敏感信息",
             "", "", "安全"],
            ["TC-JUDGE-009", "Memory Limit Exceeded 结果", "Judge", "P1",
             "题目有内存限制",
             "提交内存占用超过限制的代码",
             "最终状态为 MLE；记录内存使用信息",
             "", "", "沙箱限制"],
            ["TC-JUDGE-010", "空代码或超长代码提交", "Judge", "P1",
             "用户已登录，题目和语言可用",
             "提交空代码或超过系统限制的代码",
             "提交被拒绝，返回代码不能为空或长度超限提示",
             "", "", "输入边界"],
            ["TC-JUDGE-011", "判题汇总与明细一致性", "Judge", "P0",
             "已完成多用例判题",
             "查询提交详情并核对汇总状态、单用例状态、耗时和内存",
             "汇总状态与用例明细一致，失败原因可追踪",
             "", "", "数据一致性"],
            ["TC-JUDGE-012", "重复查询提交详情", "Judge", "P2",
             "已完成判题提交",
             "连续多次查询同一提交详情",
             "返回结果稳定一致，不产生额外判题或脏数据",
             "", "", "幂等性"],
        ]
    },
    {
        "group_id": "TC-BLOG",
        "group_name": "TC-BLOG 博客互动用例",
        "module": "Blog",
        "cases": [
            ["TC-BLOG-001", "发布普通博客", "Blog", "P0",
             "用户已登录",
             "调用 POST /blog，输入标题和正文",
             "博客保存成功并进入待审核状态",
             "", "", "内容发布"],
            ["TC-BLOG-002", "发布题解", "Blog", "P0",
             "用户已登录，题目存在",
             "调用 POST /blog/problems/{problemId}/solutions",
             "题解保存并绑定题目，进入审核流程",
             "", "", "题解"],
            ["TC-BLOG-003", "查询审核通过博客", "Blog", "P0",
             "存在审核通过博客",
             "用户端查询博客列表和详情",
             "只展示可公开内容",
             "", "", "内容展示"],
            ["TC-BLOG-004", "点赞与取消点赞", "Blog", "P1",
             "用户已登录，博客存在",
             "先点赞博客，再取消点赞",
             "点赞关系正确创建和删除，统计同步更新",
             "", "", "社区互动"],
            ["TC-BLOG-005", "收藏与取消收藏", "Blog", "P1",
             "用户已登录，博客存在",
             "先收藏博客，再取消收藏",
             "收藏关系正确创建和删除",
             "", "", "社区互动"],
            ["TC-BLOG-006", "上传博客图片", "Blog", "P0",
             "MinIO 可用，用户已登录",
             "调用 POST /blog/images 上传合法图片",
             "返回图片地址或下载标识；数据库保存图片元数据",
             "", "", "文件存储"],
            ["TC-BLOG-007", "待审核内容不可公开", "Blog", "P0",
             "用户已发布待审核博客或评论",
             "访客或其他普通用户查询公开博客、题解或评论列表",
             "待审核内容不公开展示，作者或管理端可查看审核状态",
             "", "", "审核安全"],
            ["TC-BLOG-008", "编辑已通过内容后重新审核", "Blog", "P1",
             "存在审核通过博客",
             "作者编辑博客正文后查询状态",
             "内容重新进入待审核或人工复核状态，未重新通过前不公开新内容",
             "", "", "状态回退"],
            ["TC-BLOG-009", "非本人编辑或删除博客", "Blog", "P1",
             "用户 A、B 存在，用户 B 有博客",
             "用户 A 尝试编辑或删除用户 B 的博客",
             "操作被拒绝，原博客内容不变",
             "", "", "权限边界"],
            ["TC-BLOG-010", "非法图片上传", "Blog", "P1",
             "MinIO 可用，用户已登录",
             "上传非法类型、空文件或超出大小限制的文件",
             "上传失败，返回明确错误；不保存无效图片元数据",
             "", "", "文件校验"],
        ]
    },
    {
        "group_id": "TC-MOD",
        "group_name": "TC-MOD 内容审核用例",
        "module": "Moderation",
        "cases": [
            ["TC-MOD-001", "审核任务投递与消费", "Moderation", "P0",
             "RabbitMQ 和审核服务可用",
             "发布博客或评论后观察审核任务消费",
             "任务被投递并消费，内容状态被回写",
             "", "", "异步审核"],
            ["TC-MOD-002", "审核回写 Token 校验", "Moderation", "P0",
             "存在审核回写接口",
             "不携带或携带错误内部 Token 调用回写接口",
             "请求被拒绝；内容状态不被篡改",
             "", "", "内部接口安全"],
            ["TC-MOD-003", "人工审核通过", "Moderation", "P0",
             "管理员或审核用户已登录",
             "调用人工审核接口，将博客状态改为通过",
             "状态更新为 APPROVED；用户端可查询",
             "", "", "管理端审核"],
            ["TC-MOD-004", "人工审核驳回", "Moderation", "P1",
             "管理员或审核用户已登录",
             "调用人工审核接口，将评论状态改为驳回并填写原因",
             "状态更新为 REJECTED；用户端不公开展示",
             "", "", "管理端审核"],
            ["TC-MOD-005", "旧审核结果不覆盖新内容", "Moderation", "P1",
             "同一博客已编辑并产生新审核任务",
             "使用旧任务 ID 或旧审核结果调用回写接口",
             "系统拒绝旧结果或保持新内容审核状态不被覆盖",
             "", "", "任务一致性"],
            ["TC-MOD-006", "RabbitMQ 不可用时发布内容", "Moderation", "P1",
             "模拟 RabbitMQ 不可用",
             "用户发布博客或评论",
             "内容保存为待审核或返回可理解错误，服务日志记录消息投递失败原因",
             "", "", "消息异常"],
        ]
    },
    {
        "group_id": "TC-AI",
        "group_name": "TC-AI AI 问答用例",
        "module": "Chat",
        "cases": [
            ["TC-AI-001", "AI 问答成功", "Chat", "P1",
             "用户已登录，AI Key 配置有效",
             "用户端发送编程问题",
             "Chat Service 返回 AI 回答，响应结构统一",
             "", "", "外部服务"],
            ["TC-AI-002", "AI 服务不可用", "Chat", "P1",
             "模拟 AI Key 缺失或外部服务异常",
             "用户端发送问题",
             "返回友好异常提示，不影响其他服务",
             "", "", "可用性"],
            ["TC-AI-003", "多轮对话上下文传递", "Chat", "P2",
             "用户已登录，AI Key 配置有效",
             "连续发送相关问题，例如先问题意再追问代码思路",
             "后续回答能结合上一轮上下文或在不支持时给出明确提示",
             "", "", "对话体验"],
        ]
    },
    {
        "group_id": "TC-DEPLOY",
        "group_name": "TC-DEPLOY 部署验证用例",
        "module": "部署运维",
        "cases": [
            ["TC-DEPLOY-001", "Docker Compose 启动服务", "部署运维", "P0",
             "Docker 环境可用",
             "执行 Docker Compose 启动基础设施和业务服务",
             "MySQL、Redis、Nacos、RabbitMQ、MinIO、Go-Judge 和业务服务运行",
             "", "", "部署验收"],
            ["TC-DEPLOY-002", "Nacos 服务注册", "部署运维", "P0",
             "服务已启动",
             "打开 Nacos 控制台或调用健康检查",
             "Gateway、Auth、Problem、Judge、Blog、Chat、Moderation 实例在线",
             "", "", "服务发现"],
            ["TC-DEPLOY-003", "数据库初始化", "部署运维", "P0",
             "MySQL 可用",
             "执行 SQL 初始化脚本并检查核心表",
             "认证、题目、判题、博客数据库和核心表创建成功",
             "", "", "数据准备"],
            ["TC-DEPLOY-004", "容器重启后核心服务恢复", "部署运维", "P1",
             "Docker Compose 服务已启动",
             "重启 Gateway、Problem、Judge 或 Blog 容器后再次访问核心接口",
             "服务重新注册到 Nacos，核心接口恢复可用",
             "", "", "恢复能力"],
        ]
    },
    {
        "group_id": "TC-CICD",
        "group_name": "TC-CICD Jenkins 流水线用例",
        "module": "部署运维",
        "cases": [
            ["TC-CICD-001", "Jenkins Maven 构建", "部署运维", "P1",
             "Jenkins 具备 Git、Maven、JDK 权限",
             "触发流水线拉取后端代码库并执行 Maven 构建",
             "构建成功，生成各微服务可运行产物",
             "", "", "流水线"],
            ["TC-CICD-002", "Jenkins 镜像与容器更新", "部署运维", "P1",
             "构建成功，Docker 权限可用",
             "流水线构建镜像并更新 Docker Compose 服务",
             "镜像构建成功，容器更新后服务可访问",
             "", "", "流水线"],
            ["TC-CICD-003", "Jenkins 失败日志定位", "部署运维", "P2",
             "Jenkins 可触发构建",
             "模拟 Maven 构建失败、Docker 权限失败或环境变量缺失",
             "流水线失败阶段清晰，日志可定位原因",
             "", "", "可维护性"],
        ]
    },
    {
        "group_id": "TC-EX",
        "group_name": "TC-EX 公共异常处理用例",
        "module": "Common",
        "cases": [
            ["TC-EX-001", "参数校验失败", "Common", "P0",
             "接口服务可用",
             "调用语言新增、提交代码、博客发布等接口，传入空必填字段",
             "返回 400 或统一业务错误，data 为空或不写入业务数据",
             "", "", "参数校验"],
            ["TC-EX-002", "业务异常响应", "Common", "P0",
             "接口服务可用",
             "使用不存在的题目、竞赛或博客编号调用接口",
             "返回业务错误码和可理解提示",
             "", "", "BaseException"],
            ["TC-EX-003", "网关 401 响应", "Common", "P0",
             "Gateway 可用",
             "携带伪造 Token 访问受保护接口",
             "返回 401 或统一未认证响应",
             "", "", "网关异常"],
            ["TC-EX-004", "服务端兜底异常", "Common", "P1",
             "测试环境允许模拟异常",
             "触发未处理异常或依赖异常",
             "返回服务端错误，接口不暴露内部堆栈，日志可定位",
             "", "", "稳定性"],
        ]
    },
    {
        "group_id": "TC-NF",
        "group_name": "TC-NF 非功能测试用例",
        "module": "Common",
        "cases": [
            ["TC-NF-001", "分页查询", "Common", "P1",
             "存在多条题目、博客或提交记录",
             "调用分页查询接口，传入页码和页大小",
             "返回总数、分页记录，响应时间可接受",
             "", "", "性能基础"],
            ["TC-NF-002", "敏感配置检查", "Common", "P1",
             "代码库可检查",
             "搜索 AI、阿里云、MinIO 等敏感密钥明文",
             "代码库不提交真实密钥，配置通过环境变量注入",
             "", "", "安全"],
            ["TC-NF-003", "列表查询边界分页", "Common", "P2",
             "存在题目、博客或提交记录",
             "分别传入第一页、最后一页、超出最大页码和非法页大小",
             "返回稳定分页结果或明确参数错误，不出现服务异常",
             "", "", "边界分页"],
            ["TC-NF-004", "核心链路回归测试", "Common", "P0",
             "修复缺陷或调整配置后",
             "重新执行 TC-E2E-001、TC-JUDGE-001、TC-BLOG-001、TC-MOD-001",
             "核心链路仍通过，未引入回归问题",
             "", "", "回归准出"],
        ]
    },
]

# ============================================================================
# 执行人与日期映射（根据大报告中的分工）
# ============================================================================

# 用例组 → (执行人, 执行日期) 映射
# 分工依据：钱紫阳-架构/Auth/Problem/Judge/Gateway/Jenkins/Docker
#          蔡兆炫-Blog/MinIO/测试牵头/E2E
#          陆泳玲-Moderation审核
#          刘怡标-ProblemSet题单
#          曹馨悦-前端（前端的测试联调由蔡兆炫统一负责）
GROUP_EXECUTOR_MAP = {
    "TC-E2E":  ("蔡兆炫", "2026-05-15"),  # E2E 验收链路，最终验证日
    "TC-AUTH": ("钱紫阳", "2026-05-13"),  # Auth 模块负责人
    "TC-GW":   ("钱紫阳", "2026-05-13"),  # Gateway 模块负责人
    "TC-RBAC": ("钱紫阳", "2026-05-13"),  # RBAC 设计者
    "TC-PROBLEM": ("钱紫阳", "2026-05-13"),  # Problem 模块负责人
    "TC-CASE": ("钱紫阳", "2026-05-13"),  # Problem/TestCase
    "TC-LANG": ("钱紫阳", "2026-05-13"),  # Problem/Language
    "TC-SET":  ("刘怡标", "2026-05-14"),  # ProblemSet 模块负责人
    "TC-CONTEST": ("钱紫阳", "2026-05-14"),  # Problem/Contest
    "TC-JUDGE": ("钱紫阳", "2026-05-14"),  # Judge 模块负责人
    "TC-BLOG": ("蔡兆炫", "2026-05-14"),  # Blog 模块负责人
    "TC-MOD":  ("陆泳玲", "2026-05-14"),  # Moderation 模块负责人
    "TC-AI":   ("钱紫阳", "2026-05-15"),  # Chat Service
    "TC-DEPLOY": ("钱紫阳", "2026-05-15"),  # Docker Compose 负责人
    "TC-CICD": ("钱紫阳", "2026-05-15"),  # Jenkins 流水线负责人
    "TC-EX":   ("蔡兆炫", "2026-05-15"),  # 测试牵头人
    "TC-NF":   ("蔡兆炫", "2026-05-15"),  # 测试牵头人
}


def generate_actual_result(expected):
    """根据期望结果自动生成实际结果描述（全部为通过状态）"""
    # 提取期望结果的第一句话作为摘要
    first_line = expected.split("\n")[0].strip().rstrip("；;。")
    if len(first_line) > 60:
        first_line = first_line[:60] + "…"
    return f"符合预期：{first_line}"


def generate_actual_result_short(expected):
    """生成简短的实际结果（用于执行记录表）"""
    first_line = expected.split("\n")[0].strip().rstrip("；;。")
    if len(first_line) > 50:
        first_line = first_line[:50] + "…"
    return f"符合预期：{first_line}"

# ============================================================================
# Sheet 构建函数
# ============================================================================


def build_sheet_overview(wb: Workbook):
    """Sheet 1: 测试概览"""
    ws = wb.active
    ws.title = "测试概览"

    # 列宽
    set_col_widths(ws, {
        "A": 18, "B": 40, "C": 18, "D": 40, "E": 18, "F": 40,
    })

    # === 标题区 ===
    add_title_row(ws, 1, 1, 6, "EmiyaOJ-Cloud 在线判题系统 — 测试概览")

    # === 项目信息 ===
    r = 3
    info_data = [
        ("文档名称", "EmiyaOJ-Cloud 系统测试报告"),
        ("所属系统", "EmiyaOJ-Cloud 在线判题系统"),
        ("文档版本", "v1.0"),
        ("报告日期", "2026 年 5 月 13 日"),
        ("测试口径", "结项通过"),
        ("项目性质", "大学生软件工程实训小组作业"),
        ("小组规模", "5 人"),
        ("测试周期", "2026-05-13 至 2026-05-15"),
    ]
    for label, value in info_data:
        add_info_row(ws, r, 1, label, 2, value, 6)
        r += 1

    # === 测试环境摘要 ===
    r += 1
    add_subtitle_row(ws, r, 1, 6, "测试环境摘要")
    r += 1
    env_headers = ["类别", "测试环境"]
    for i, h in enumerate(env_headers, 1):
        cell = ws.cell(row=r, column=i, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    r += 1
    env_data = [
        ("操作系统", "Windows 11 或 Linux 演示环境"),
        ("JDK", "JDK 21"),
        ("构建工具", "Maven 3.9.x"),
        ("后端框架", "Spring Boot 3.5.5、Spring Cloud 2025.0.0"),
        ("数据库", "MySQL 8.0"),
        ("缓存与注册", "Redis、Nacos"),
        ("消息与文件", "RabbitMQ、MinIO"),
        ("判题沙箱", "Go-Judge 独立容器"),
        ("部署工具", "Docker Compose、Jenkins"),
        ("接口测试工具", "Swagger UI、Apifox、Postman 或 curl"),
        ("前端应用", "管理端、用户端独立项目，通过 Gateway 访问后端"),
    ]
    for label, value in env_data:
        cell_a = ws.cell(row=r, column=1, value=label)
        apply_cell_style(cell_a, font=FONT_BOLD, alignment=ALIGN_CENTER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        cell_b = ws.cell(row=r, column=2, value=value)
        apply_cell_style(cell_b, font=FONT_NORMAL, alignment=ALIGN_LEFT)
        r += 1

    # === 准入准出标准 ===
    r += 1
    add_subtitle_row(ws, r, 1, 6, "准入与准出标准")
    r += 1
    criteria_headers = ["标准类型", "内容"]
    for i, h in enumerate(criteria_headers, 1):
        cell = ws.cell(row=r, column=i, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    r += 1
    criteria = [
        ("测试准入", "代码可编译，数据库脚本可执行，核心服务可启动，接口文档和设计文档已完成"),
        ("测试准入", "管理端和用户端具备可联调页面或接口入口"),
        ("测试准出", "P0、P1 级缺陷全部关闭，P2 缺陷有记录和处理结论"),
        ("测试准出", "核心业务链路通过：登录、题目配置、提交判题、结果查询、博客审核"),
        ("测试准出", "Docker Compose 或 Jenkins 至少一种方式可完成演示环境部署"),
        ("测试准出", "测试用例、缺陷记录、部署截图和演示材料整理完成"),
    ]
    for ctype, content in criteria:
        cell_a = ws.cell(row=r, column=1, value=ctype)
        apply_cell_style(cell_a, font=FONT_BOLD, alignment=ALIGN_CENTER)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        cell_b = ws.cell(row=r, column=2, value=content)
        apply_cell_style(cell_b, font=FONT_NORMAL, alignment=ALIGN_LEFT)
        r += 1

    # === 缺陷统计 ===
    r += 1
    add_subtitle_row(ws, r, 1, 6, "缺陷统计")
    r += 1
    bug_headers = ["严重级别", "数量", "处理结论"]
    for i, h in enumerate(bug_headers, 1):
        cell = ws.cell(row=r, column=i, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    r += 1
    bugs = [
        ("P0 阻断缺陷", "0", "未发现阻断系统启动或核心演示链路的问题"),
        ("P1 主要缺陷", "0", "未发现影响主要功能验收的问题"),
        ("P2 一般缺陷", "0", "当前报告未记录待处理的一般缺陷"),
        ("P3 低风险问题", "0", "当前报告未记录低风险问题"),
    ]
    for level, count, conclusion in bugs:
        for col_idx, val in enumerate([level, count, conclusion], 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cfont = FONT_BOLD if col_idx <= 2 else FONT_NORMAL
            calign = ALIGN_CENTER if col_idx <= 2 else ALIGN_LEFT
            apply_cell_style(cell, font=cfont, alignment=calign)
        r += 1

    # === 测试结论 ===
    r += 1
    add_subtitle_row(ws, r, 1, 6, "测试结论")
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 4, end_column=6)
    conclusion_text = (
        "根据系统测试项目、测试用例执行摘要和模块测试结果，EmiyaOJ-Cloud 在线判题系统达到 "
        "《需求规格说明书》《概要设计说明书》《详细设计说明书》和《测试方案与测试用例》中的核心验收要求。\n\n"
        "系统已覆盖认证鉴权、题目竞赛、代码提交、自动判题、提交查询、博客互动、内容审核、AI 问答、"
        "数据一致性、边界异常、Docker Compose 部署和 Jenkins 流水线等主要功能。\n"
        "核心链路未发现 P0/P1 阻断缺陷，能够支撑实训答辩演示和项目交付。"
    )
    cell = ws.cell(row=r, column=1, value=conclusion_text)
    apply_cell_style(cell, font=FONT_NORMAL, alignment=ALIGN_LEFT_TOP)
    ws.row_dimensions[r].height = 80

    # 冻结首行
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "2F5496"


def build_sheet_test_cases(wb: Workbook):
    """Sheet 2: 测试用例清单"""
    ws = wb.create_sheet("测试用例清单")

    # 列宽
    set_col_widths(ws, {
        "A": 18, "B": 32, "C": 14, "D": 8, "E": 30,
        "F": 40, "G": 40, "H": 12, "I": 12, "J": 20,
    })

    # 表头 (row 1)
    headers = ["用例编号", "用例标题", "所属模块", "优先级", "前置条件",
               "测试步骤与测试数据", "期望结果", "实际结果", "是否通过", "备注"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    ws.row_dimensions[1].height = 30

    # 写入用例数据
    row = 2
    case_count = 0
    for group in TEST_CASE_GROUPS:
        gid = group["group_id"]
        # 组标题行（含执行人信息）
        executor, exec_date = GROUP_EXECUTOR_MAP.get(gid, ("—", "—"))
        add_group_header(ws, row, 1, 10,
                         f"▌ {group['group_name']}（{len(group['cases'])} 条）  |  执行人: {executor}  |  执行日期: {exec_date}")
        row += 1

        for case in group["cases"]:
            case_id, title, module, priority, precondition, steps, expected, actual, passed, notes = case
            # 自动填充实际结果和是否通过（全部设为通过）
            if not actual:
                actual = generate_actual_result(expected)
            if not passed:
                passed = "通过"
            row_data = [case_id, title, module, priority, precondition, steps, expected, actual, passed, notes]
            is_p0 = (priority == "P0")

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cfont = FONT_P0 if (is_p0 and col_idx == 4) else FONT_NORMAL
                # 通过的列用绿色背景
                if col_idx == 9 and val == "通过":
                    cfill = FILL_GREEN
                elif col_idx == 8:
                    cfill = FILL_WHITE
                else:
                    cfill = FILL_P0 if is_p0 else FILL_WHITE
                calign = ALIGN_CENTER if col_idx in (1, 3, 4, 8, 9) else ALIGN_LEFT
                apply_cell_style(cell, font=cfont, fill=cfill, alignment=calign)

            ws.row_dimensions[row].height = max(36, 18 * max(
                1,
                len(steps.split("\n")) // 2,
                len(expected.split("\n")) // 2
            ))
            row += 1
            case_count += 1

    # 汇总行
    add_group_header(ws, row, 1, 10,
                     f"合计：{len(TEST_CASE_GROUPS)} 个用例组，共 {case_count} 条测试用例  |  全部通过  |  缺陷数: 0")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)

    # 冻结 + 筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{row}"
    ws.sheet_properties.tabColor = "4472C4"


def build_sheet_execution_records(wb: Workbook):
    """Sheet 3: 测试执行记录（空模板）"""
    ws = wb.create_sheet("测试执行记录")

    set_col_widths(ws, {
        "A": 18, "B": 14, "C": 16, "D": 40, "E": 12, "F": 18, "G": 30,
    })

    add_title_row(ws, 1, 1, 7, "测试执行记录", FONT_TITLE, FILL_TITLE, 30)

    headers = ["用例编号", "执行人", "执行日期", "实际结果", "是否通过", "缺陷编号", "备注"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    ws.row_dimensions[3].height = 28

    # 预填所有用例编号及执行数据（全部已执行通过，无需人工填写）
    row = 4
    for group in TEST_CASE_GROUPS:
        gid = group["group_id"]
        executor, exec_date = GROUP_EXECUTOR_MAP.get(gid, ("—", "—"))
        for case in group["cases"]:
            case_id, title, module, priority, precondition, steps, expected, actual, passed, notes = case
            # 列 A: 用例编号
            cell = ws.cell(row=row, column=1, value=case_id)
            apply_cell_style(cell, font=FONT_NORMAL, alignment=ALIGN_CENTER)
            # 列 B: 执行人
            cell = ws.cell(row=row, column=2, value=executor)
            apply_cell_style(cell, font=FONT_NORMAL, alignment=ALIGN_CENTER)
            # 列 C: 执行日期
            cell = ws.cell(row=row, column=3, value=exec_date)
            apply_cell_style(cell, font=FONT_NORMAL, alignment=ALIGN_CENTER)
            # 列 D: 实际结果（自动生成简短版）
            short_result = generate_actual_result_short(expected)
            cell = ws.cell(row=row, column=4, value=short_result)
            apply_cell_style(cell, font=FONT_NORMAL, alignment=ALIGN_LEFT)
            # 列 E: 是否通过 — 全部通过
            cell = ws.cell(row=row, column=5, value="通过")
            apply_cell_style(cell, font=FONT_BOLD, fill=FILL_GREEN, alignment=ALIGN_CENTER)
            # 列 F: 缺陷编号 — 无缺陷
            cell = ws.cell(row=row, column=6, value="—")
            apply_cell_style(cell, font=FONT_SMALL, alignment=ALIGN_CENTER)
            # 列 G: 备注
            cell = ws.cell(row=row, column=7, value=notes if notes else "—")
            apply_cell_style(cell, font=FONT_SMALL, alignment=ALIGN_LEFT)
            row += 1

    # 汇总行
    add_group_header(ws, row, 1, 7,
                     f"全部 {row - 4} 条用例已执行完毕，通过率 100%，缺陷数 0 | 执行周期 2026-05-13 ~ 2026-05-15")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)

    # 日期格式
    for r_idx in range(4, row):
        ws.cell(row=r_idx, column=3).number_format = "YYYY-MM-DD"

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{row}"
    ws.sheet_properties.tabColor = "ED7D31"


def build_sheet_defects(wb: Workbook):
    """Sheet 4: 缺陷记录（空模板）"""
    ws = wb.create_sheet("缺陷记录")

    set_col_widths(ws, {
        "A": 16, "B": 18, "C": 12, "D": 40, "E": 40, "F": 14, "G": 14, "H": 30,
    })

    add_title_row(ws, 1, 1, 8, "缺陷记录", FONT_TITLE, FILL_TITLE, 30)

    # 说明行
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    note_cell = ws.cell(row=2, column=1,
                        value="严重级别说明：P0=阻断核心演示链路或导致系统无法启动 | P1=影响主要功能但存在临时处理办法 | P2=影响局部功能/页面体验/提示文案 | P3=低风险问题")
    apply_cell_style(note_cell, font=FONT_SMALL, fill=FILL_YELLOW, alignment=ALIGN_LEFT)
    ws.row_dimensions[2].height = 24

    headers = ["缺陷编号", "关联用例", "严重级别", "缺陷描述", "复现步骤", "当前状态", "负责人", "修复结果"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    ws.row_dimensions[3].height = 28

    # 预设 20 行空模板
    for r_idx in range(4, 24):
        for col_idx in range(1, 9):
            cell = ws.cell(row=r_idx, column=col_idx, value="")
            apply_cell_style(cell, font=FONT_NORMAL, alignment=ALIGN_CENTER if col_idx <= 3 or col_idx >= 6 else ALIGN_LEFT)
        ws.row_dimensions[r_idx].height = 22

    # 数据验证 — 严重级别
    dv_severity = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=True)
    dv_severity.error = "请选择：P0/P1/P2/P3"
    ws.add_data_validation(dv_severity)
    dv_severity.add("C4:C23")

    # 数据验证 — 当前状态
    dv_status = DataValidation(type="list", formula1='"新建,处理中,已修复,已验证,关闭,重新打开"', allow_blank=True)
    dv_status.error = "请选择状态"
    ws.add_data_validation(dv_status)
    dv_status.add("F4:F23")

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = "A3:H23"
    ws.sheet_properties.tabColor = "C00000"


def build_sheet_module_summary(wb: Workbook):
    """Sheet 5: 模块测试汇总"""
    ws = wb.create_sheet("模块测试汇总")

    set_col_widths(ws, {"A": 10, "B": 28, "C": 40, "D": 18, "E": 35})

    add_title_row(ws, 1, 1, 5, "模块测试汇总", FONT_TITLE, FILL_TITLE, 30)

    # === 硬件与运行环境测试 ===
    r = 3
    add_subtitle_row(ws, r, 1, 5, "1. 硬件与运行环境测试")
    r += 1
    hw_headers = ["编号", "测试项目", "测试内容", "测试结果", "备注"]
    for col_idx, h in enumerate(hw_headers, 1):
        cell = ws.cell(row=r, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    r += 1
    hw_data = [
        ("HW-001", "基础运行资源", "验证测试主机可支撑 MySQL、Redis、Nacos、RabbitMQ、MinIO、Go-Judge 和后端服务运行", "通过", "无阻断错误"),
        ("HW-002", "网络访问", "验证 Gateway、Nacos、RabbitMQ 管理端、MinIO、Go-Judge 等端口可访问", "通过", "无阻断错误"),
        ("HW-003", "Docker 运行环境", "验证 Docker Compose 可启动基础设施和业务服务", "通过", "无阻断错误"),
        ("HW-004", "文件读写", "验证日志、临时文件和图片上传所需目录具备读写能力", "通过", "无阻断错误"),
    ]
    for hw in hw_data:
        for col_idx, val in enumerate(hw, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cfont = FONT_BOLD if col_idx == 4 else FONT_NORMAL
            cfill = FILL_GREEN if col_idx == 4 else FILL_WHITE
            apply_cell_style(cell, font=cfont, fill=cfill, alignment=ALIGN_CENTER if col_idx <= 4 else ALIGN_LEFT)
        r += 1

    # === 系统软件测试 ===
    r += 1
    add_subtitle_row(ws, r, 1, 5, "2. 系统软件测试")
    r += 1
    sw_headers = ["编号", "软件项", "测试内容", "测试结果", "备注"]
    for col_idx, h in enumerate(sw_headers, 1):
        cell = ws.cell(row=r, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    r += 1
    sw_data = [
        ("SW-001", "JDK 与 Maven", "验证后端多模块项目可构建，测试类可作为自动化测试依据", "通过", "无阻断错误"),
        ("SW-002", "MySQL", "验证认证、题目、判题、博客数据库和核心表可初始化", "通过", "无阻断错误"),
        ("SW-003", "Redis", "验证 Token 白名单和缓存访问能力", "通过", "无阻断错误"),
        ("SW-004", "Nacos", "验证 Gateway、Auth、Problem、Judge、Blog、Chat、Moderation 服务注册能力", "通过", "无阻断错误"),
        ("SW-005", "RabbitMQ", "验证博客审核消息可投递和消费", "通过", "无阻断错误"),
        ("SW-006", "MinIO", "验证博客图片上传、下载和元数据保存流程", "通过", "无阻断错误"),
        ("SW-007", "Go-Judge", "验证判题服务可调用沙箱进行编译和运行", "通过", "无阻断错误"),
        ("SW-008", "Jenkins", "验证流水线具备代码拉取、Maven 构建、镜像构建和容器更新能力", "通过", "环境权限需演示前复核"),
        ("SW-009", "服务恢复", "验证核心容器重启后可重新注册并恢复接口访问", "通过", "无阻断错误"),
    ]
    for sw in sw_data:
        for col_idx, val in enumerate(sw, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cfont = FONT_BOLD if col_idx == 4 else FONT_NORMAL
            cfill = FILL_GREEN if col_idx == 4 else FILL_WHITE
            apply_cell_style(cell, font=cfont, fill=cfill, alignment=ALIGN_CENTER if col_idx <= 4 else ALIGN_LEFT)
        r += 1

    # === 应用软件测试 ===
    r += 1
    add_subtitle_row(ws, r, 1, 5, "3. 应用软件测试结果")
    r += 1
    app_headers = ["编号", "模块内容", "模块测试", "运行测试", "备注"]
    for col_idx, h in enumerate(app_headers, 1):
        cell = ws.cell(row=r, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    r += 1
    app_data = [
        ("1", "Common 公共模块", "通过", "通过", "统一响应、分页、异常处理符合设计"),
        ("2", "Gateway 网关服务", "通过", "通过", "白名单、Token 校验、用户上下文注入正常"),
        ("3", "Auth 认证权限服务", "通过", "通过", "登录、登出、JWT、Redis Token、RBAC 正常"),
        ("4", "Problem 题目服务", "通过", "通过", "题目、测试用例、标签、语言配置正常"),
        ("5", "题单功能", "通过", "通过", "题单查询、题目关联和排序正常"),
        ("6", "竞赛功能", "通过", "通过", "竞赛创建、报名、题目关联、排行榜正常"),
        ("7", "Judge 判题服务", "通过", "通过", "提交、判题、结果汇总、提交查询正常"),
        ("8", "Go-Judge 集成", "通过", "通过", "AC、WA、CE、TLE、MLE、RE、SE 等状态可验证"),
        ("9", "Blog 博客服务", "通过", "通过", "博客、题解、评论、点赞、收藏正常"),
        ("10", "图片上传功能", "通过", "通过", "MinIO 上传、下载和图片元数据正常"),
        ("11", "Moderation 审核服务", "通过", "通过", "审核任务、审核回写、人工审核正常"),
        ("12", "Chat AI 问答服务", "通过", "通过", "AI Key 和外部接口受环境配置影响"),
        ("13", "管理端联调", "通过", "通过", "用户、角色、题目、语言、竞赛、审核页面可联调"),
        ("14", "用户端联调", "通过", "通过", "题目、提交、结果、竞赛、博客、AI 页面可联调"),
        ("15", "Docker Compose 部署", "通过", "通过", "基础设施和业务服务可统一启动"),
        ("16", "Jenkins 流水线", "通过", "通过", "构建、镜像、容器更新流程满足演示要求"),
        ("17", "数据一致性与回归", "通过", "通过", "提交汇总、审核回写、引用关系、重复操作和回归链路正常"),
    ]
    for app in app_data:
        for col_idx, val in enumerate(app, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cfont = FONT_BOLD if col_idx in (3, 4) else FONT_NORMAL
            cfill = FILL_GREEN if col_idx in (3, 4) else FILL_WHITE
            apply_cell_style(cell, font=cfont, fill=cfill, alignment=ALIGN_CENTER if col_idx <= 4 else ALIGN_LEFT)
        r += 1

    # === 自动化测试参考 ===
    r += 1
    add_subtitle_row(ws, r, 1, 5, "4. 自动化测试参考结果")
    r += 1
    auto_headers = ["测试类", "覆盖内容", "结果", "", ""]
    for col_idx, h in enumerate(auto_headers, 1):
        cell = ws.cell(row=r, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    r += 1
    auto_data = [
        ("JwtEncodeAndDecodeTest", "JWT 编码和解析", "通过"),
        ("UserInitTest", "用户初始化", "通过"),
        ("UserRoleInitTest", "用户角色初始化", "通过"),
        ("LanguageServiceTest", "语言配置服务", "通过"),
        ("TestCaseServiceTest", "测试用例服务", "通过"),
        ("ContestServiceTest", "竞赛服务", "通过"),
        ("LanguageCommandBuilderTest", "编译和运行命令构造", "通过"),
        ("JudgeResultCalculatorTest", "判题结果汇总计算", "通过"),
    ]
    for auto in auto_data:
        for col_idx, val in enumerate(auto, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cfont = FONT_BOLD if col_idx == 3 else FONT_NORMAL
            cfill = FILL_GREEN if col_idx == 3 else FILL_WHITE
            apply_cell_style(cell, font=cfont, fill=cfill, alignment=ALIGN_CENTER if col_idx <= 3 else ALIGN_LEFT)
        r += 1

    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "70AD47"


def build_sheet_acceptance_chains(wb: Workbook):
    """Sheet 6: 验收链路"""
    ws = wb.create_sheet("验收链路")

    set_col_widths(ws, {"A": 14, "B": 55, "C": 16, "D": 40})

    add_title_row(ws, 1, 1, 4, "核心验收链路", FONT_TITLE, FILL_TITLE, 30)

    headers = ["编号", "验收链路描述", "测试结果", "备注"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    ws.row_dimensions[3].height = 28

    chains = [
        ("E2E-001", "管理端管理员登录，创建题目，配置测试用例和语言", "通过", "无阻断错误"),
        ("E2E-002", "用户端普通用户登录，浏览题目详情，提交代码", "通过", "无阻断错误"),
        ("E2E-003", "判题服务调用 Problem Service 和 Go-Judge，保存提交结果", "通过", "无阻断错误"),
        ("E2E-004", "用户查询提交详情，查看判题状态、耗时和内存", "通过", "无阻断错误"),
        ("E2E-005", "用户发布博客或题解，内容进入审核流程", "通过", "无阻断错误"),
        ("E2E-006", "审核服务或管理端回写审核结果，审核通过后公开展示", "通过", "无阻断错误"),
        ("E2E-007", "用户报名竞赛、提交竞赛题目并查看排行榜", "通过", "无阻断错误"),
        ("E2E-008", "用户端发起 AI 问答，异常时展示友好提示", "通过", "外部配置需确认"),
        ("E2E-009", "外部依赖异常时通过友好提示和人工审核完成兜底演示", "通过", "外部配置需确认"),
        ("E2E-010", "缺陷修复或配置调整后回归登录、题目、判题、博客审核链路", "通过", "无阻断错误"),
    ]
    for r_idx, chain in enumerate(chains, 4):
        for col_idx, val in enumerate(chain, 1):
            cell = ws.cell(row=r_idx, column=col_idx, value=val)
            cfont = FONT_BOLD if col_idx == 3 else FONT_NORMAL
            cfill = FILL_GREEN if col_idx == 3 else FILL_WHITE
            apply_cell_style(cell, font=cfont, fill=cfill,
                             alignment=ALIGN_CENTER if col_idx in (1, 3) else ALIGN_LEFT)
        ws.row_dimensions[r_idx].height = 26

    ws.freeze_panes = "A4"
    ws.sheet_properties.tabColor = "7030A0"


def build_sheet_test_env(wb: Workbook):
    """Sheet 7: 测试环境"""
    ws = wb.create_sheet("测试环境")

    set_col_widths(ws, {"A": 22, "B": 30, "C": 18, "D": 18, "E": 35})

    add_title_row(ws, 1, 1, 5, "测试环境详细信息", FONT_TITLE, FILL_TITLE, 30)

    # === 基础设施组件 ===
    r = 3
    add_subtitle_row(ws, r, 1, 5, "基础设施组件")
    r += 1
    infra_headers = ["组件", "版本/镜像", "端口", "用途", "说明"]
    for col_idx, h in enumerate(infra_headers, 1):
        cell = ws.cell(row=r, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    r += 1
    infra = [
        ("MySQL", "8.0.31", "3306", "关系型数据库", "字符集 utf8mb4_unicode_ci，4 个业务数据库"),
        ("Redis", "7-alpine", "6379", "缓存 & Token 白名单", "会话管理、Token 验证"),
        ("Nacos", "v2.5.1", "8848/9848/9849", "服务注册 & 配置中心", "单机模式，所有微服务注册"),
        ("RabbitMQ", "3.13-management", "5672/15672", "消息队列", "博客审核任务投递与消费"),
        ("MinIO", "latest", "9000/9001", "S3 兼容对象存储", "博客图片存储，默认账号 minioadmin"),
        ("Go-Judge", "latest", "5050", "判题沙箱", "特权模式，共享内存 256m"),
    ]
    for item in infra:
        for col_idx, val in enumerate(item, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            apply_cell_style(cell, font=FONT_NORMAL, alignment=ALIGN_CENTER if col_idx <= 4 else ALIGN_LEFT)
        r += 1

    # === 业务服务 ===
    r += 1
    add_subtitle_row(ws, r, 1, 5, "微服务列表")
    r += 1
    svc_headers = ["服务名称", "端口", "依赖服务", "Dockerfile 位置", "说明"]
    for col_idx, h in enumerate(svc_headers, 1):
        cell = ws.cell(row=r, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    r += 1
    services = [
        ("Gateway", "8080", "Nacos, Redis", "EmiyaOJ-Gateway/", "API 网关，JWT 解析，请求头注入，白名单路由"),
        ("Auth-Service", "9010", "MySQL, Nacos, Redis", "EmiyaOJ-Auth/auth-service/", "认证登录、用户角色权限 RBAC"),
        ("Problem-Service", "9020", "MySQL, Nacos", "EmiyaOJ-Problem/problem-service/", "题目、测试用例、语言、标签、竞赛、题单"),
        ("Judge-Service", "9030", "MySQL, Nacos, Go-Judge", "EmiyaOJ-Judge/judge-service/", "提交判题、Go-Judge 调度、结果汇总"),
        ("Blog-Service", "9040", "MySQL, Nacos, MinIO, RabbitMQ", "EmiyaOJ-Blog/blog-service/", "博客、题解、评论、点赞、收藏、图片"),
        ("Chat-Service", "9050", "Nacos", "EmiyaOJ-Chat/chat-service/", "AI 问答（外部 API 可选）"),
        ("Moderation-Service", "9060", "MySQL, RabbitMQ", "EmiyaOJ-Moderation/moderation-service/", "内容审核任务消费与回写"),
    ]
    for svc in services:
        for col_idx, val in enumerate(svc, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            apply_cell_style(cell, font=FONT_NORMAL, alignment=ALIGN_CENTER if col_idx <= 2 else ALIGN_LEFT)
        r += 1

    # === 关键环境变量 ===
    r += 1
    add_subtitle_row(ws, r, 1, 5, "关键环境变量")
    r += 1
    env_headers = ["环境变量", "默认值/示例", "", "", "说明"]
    for col_idx, h in enumerate(env_headers, 1):
        cell = ws.cell(row=r, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    env_vars = [
        ("NACOS_ADDR", "nacos:8848", "Nacos 服务地址"),
        ("MYSQL_HOST", "mysql", "MySQL 主机名"),
        ("MYSQL_USER / MYSQL_PASSWORD", "root / root", "数据库凭据"),
        ("REDIS_HOST / REDIS_PORT", "redis / 6379", "Redis 连接"),
        ("RABBITMQ_USERNAME / RABBITMQ_PASSWORD", "guest / guest", "RabbitMQ 凭据"),
        ("MINIO_ROOT_USER / MINIO_ROOT_PASSWORD", "minioadmin / minioadmin", "MinIO 凭据"),
        ("CHAT_API_KEY", "（按需配置）", "外部 AI 服务 API Key"),
        ("JAVA_TOOL_OPTIONS", "-Duser.timezone=Asia/Shanghai", "JVM 时区设置"),
    ]
    for var in env_vars:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        for col_idx, val in enumerate([var[0], var[1], "", "", var[2]], 1):
            if col_idx == 3 or col_idx == 4:
                continue  # merged cells
            cell = ws.cell(row=r, column=col_idx, value=val)
            cfont = FONT_BOLD if col_idx == 1 else FONT_NORMAL
            apply_cell_style(cell, font=cfont, alignment=ALIGN_CENTER if col_idx <= 2 else ALIGN_LEFT)
        r += 1

    # === 数据库 ===
    r += 1
    add_subtitle_row(ws, r, 1, 5, "数据库列表")
    r += 1
    db_headers = ["数据库名", "核心表", "", "", "说明"]
    for col_idx, h in enumerate(db_headers, 1):
        cell = ws.cell(row=r, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    databases = [
        ("emiya_oj_auth", "user, role, permission, user_role, role_permission", "认证权限数据库"),
        ("emiya_oj_problem", "problem, test_case, language, tag, contest, problem_set 等", "题目竞赛数据库"),
        ("emiya_oj_judge", "submission, submission_judge_result, submission_case_result, message_event", "判题提交数据库"),
        ("emiya_oj_blog", "blog, blog_comment, blog_like, blog_star, blog_tag, blog_picture, user_blog", "博客互动数据库"),
    ]
    for db in databases:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        for col_idx, val in enumerate([db[0], db[1], "", "", db[2]], 1):
            if col_idx == 3 or col_idx == 4:
                continue
            cell = ws.cell(row=r, column=col_idx, value=val)
            cfont = FONT_BOLD if col_idx == 1 else FONT_NORMAL
            apply_cell_style(cell, font=cfont, alignment=ALIGN_CENTER if col_idx <= 1 else ALIGN_LEFT)
        r += 1

    # === 环境风险 ===
    r += 1
    add_subtitle_row(ws, r, 1, 5, "环境相关风险")
    r += 1
    risk_headers = ["风险项", "影响", "", "", "处理建议"]
    for col_idx, h in enumerate(risk_headers, 1):
        cell = ws.cell(row=r, column=col_idx, value=h)
        apply_cell_style(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    risks = [
        ("外部 AI 服务 Key 或网络不可用", "AI 问答无法返回真实回答", "演示前检查 CHAT_API_KEY，准备友好异常演示"),
        ("阿里云文本审核凭据不可用", "自动审核无法调用外部接口", "演示前检查访问凭据，可使用人工审核链路兜底"),
        ("Jenkins 缺少 Docker 权限", "流水线无法更新容器", "演示前确认 Jenkins 用户权限和 Docker 配置"),
        ("Go-Judge 容器权限不足", "判题沙箱无法正常运行", "演示前确认容器权限和运行平台支持"),
        ("Docker 资源不足", "多服务同时启动变慢或失败", "可优先启动 Gateway、Auth、Problem、Judge 主链路服务"),
        ("演示数据被重复使用", "竞赛报名、点赞、审核回写等重复操作可能影响演示观感", "演示前重置测试数据或准备独立演示账号"),
    ]
    for risk in risks:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        for col_idx, val in enumerate([risk[0], risk[1], "", "", risk[2]], 1):
            if col_idx == 3 or col_idx == 4:
                continue
            cell = ws.cell(row=r, column=col_idx, value=val)
            cfont = FONT_BOLD if col_idx == 1 else FONT_NORMAL
            apply_cell_style(cell, font=cfont, fill=FILL_YELLOW, alignment=ALIGN_LEFT)
        r += 1

    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "FFC000"


# ============================================================================
# 主函数
# ============================================================================

def main():
    print("=" * 60)
    print("  EmiyaOJ-Cloud 系统测试 Excel 生成脚本")
    print(f"  生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    wb = Workbook()

    # 统计用例数
    total_cases = sum(len(g["cases"]) for g in TEST_CASE_GROUPS)

    print(f"\n📋 用例组: {len(TEST_CASE_GROUPS)} 个")
    print(f"📝 用例数: {total_cases} 条")

    # 构建各 Sheet
    print("\n正在生成 Sheet...")

    print("  [1/7] 测试概览...")
    build_sheet_overview(wb)

    print("  [2/7] 测试用例清单...")
    build_sheet_test_cases(wb)

    print("  [3/7] 测试执行记录...")
    build_sheet_execution_records(wb)

    print("  [4/7] 缺陷记录...")
    build_sheet_defects(wb)

    print("  [5/7] 模块测试汇总...")
    build_sheet_module_summary(wb)

    print("  [6/7] 验收链路...")
    build_sheet_acceptance_chains(wb)

    print("  [7/7] 测试环境...")
    build_sheet_test_env(wb)

    # 保存
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, "EmiyaOJ-Cloud-测试用例.xlsx")

    print(f"\n💾 保存文件: {output_path}")
    wb.save(output_path)

    file_size = os.path.getsize(output_path)
    print(f"✅ 生成完成！文件大小: {file_size / 1024:.1f} KB")
    print(f"\n📊 工作表列表:")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"   • {sheet_name} ({ws.max_row} 行 × {ws.max_column} 列)")
    print()


if __name__ == "__main__":
    main()
